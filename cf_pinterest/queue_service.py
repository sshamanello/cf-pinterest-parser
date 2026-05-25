import json
import csv
import os
from pathlib import Path

from cf_pinterest.db import (
    check_db_health,
    connect_db,
    insert_sync_run,
    load_queue_stats,
    list_publish_items_for_n8n,
    load_queue_summary,
    rebuild_publish_items,
    rebuild_publish_items_all,
    upsert_queue_items,
)
from cf_pinterest.models import QueueItem, QueueSyncResult

DEFAULT_N8N_EXPORT_PROFILES: dict[str, list[str]] = {
    "n8n_default": ["slug", "title", "description", "image_url", "target_url", "status", "updated_at"],
    "n8n_minimal": ["title", "description", "image_url", "target_url"],
}
EXPORT_PROFILES_PATH = Path(
    os.environ.get("CF_N8N_EXPORT_PROFILES_PATH", str(Path(__file__).resolve().parent / "export_profiles.json"))
)


def _load_export_profiles() -> dict[str, list[str]]:
    if not EXPORT_PROFILES_PATH.exists():
        return DEFAULT_N8N_EXPORT_PROFILES

    try:
        raw = json.loads(EXPORT_PROFILES_PATH.read_text(encoding="utf-8"))
    except Exception:
        return DEFAULT_N8N_EXPORT_PROFILES

    profiles: dict[str, list[str]] = {}
    if not isinstance(raw, dict):
        return DEFAULT_N8N_EXPORT_PROFILES

    for name, columns in raw.items():
        if not isinstance(name, str) or not isinstance(columns, list):
            continue
        normalized = [str(x).strip() for x in columns if str(x).strip()]
        if normalized:
            profiles[name] = normalized

    return profiles or DEFAULT_N8N_EXPORT_PROFILES


def get_export_profiles() -> dict[str, list[str]]:
    return _load_export_profiles()


def _row_to_queue_item(row: dict, niche: str) -> QueueItem | None:
    slug = str(row.get("slug") or "").strip()
    if not slug:
        source_file = str(row.get("source_file") or "").strip()
        slug = Path(source_file).stem if source_file else ""
    if not slug:
        return None

    return QueueItem(
        slug=slug,
        title=str(row.get("title") or ""),
        niche=niche,
        status=str(row.get("status") or ""),
        pin_jpg=str(row.get("pin_jpg") or ""),
        cf_url=str(row.get("cf_url") or ""),
        affiliate_url=str(row.get("affiliate_url") or ""),
        image_url=str(row.get("image_url") or ""),
        source_file=str(row.get("source_file") or ""),
    )


def _collect_status_stats(items: list[QueueItem]) -> tuple[int, int, int]:
    generated = sum(1 for item in items if item.status == "generated")
    uploaded = sum(1 for item in items if item.status == "uploaded")
    rejected = sum(1 for item in items if item.status == "rejected")
    return generated, uploaded, rejected


def sync_report_to_queue_db(report_path: str | Path, db_path: str | Path, niche: str) -> QueueSyncResult:
    raw = json.loads(Path(report_path).read_text(encoding="utf-8"))
    queue_items: list[QueueItem] = []
    skipped = 0
    for item in raw.get("items", []):
        mapped = _row_to_queue_item(item, niche=niche)
        if mapped is None:
            skipped += 1
            continue
        queue_items.append(mapped)

    generated, uploaded, rejected = _collect_status_stats(queue_items)
    result = QueueSyncResult(
        parsed_items=len(raw.get("items", [])),
        upserted_items=len(queue_items),
        skipped_items=skipped,
        generated_items=generated,
        uploaded_items=uploaded,
        rejected_items=rejected,
    )

    with connect_db(db_path) as conn:
        upsert_queue_items(conn, queue_items)
        rebuild_publish_items(conn, niche)
        insert_sync_run(conn, str(report_path), niche, result)
    return result


def import_sheet_file_to_queue_db(file_path: str | Path, db_path: str | Path, niche: str) -> QueueSyncResult:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    rows: list[dict] = []
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required to import .xlsx files") from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if values:
            headers = [str(x).strip() if x is not None else "" for x in values[0]]
            for row_values in values[1:]:
                row = {
                    headers[idx]: (row_values[idx] if idx < len(row_values) else "")
                    for idx in range(len(headers))
                    if headers[idx]
                }
                rows.append(row)
        wb.close()
    else:
        raise ValueError("Unsupported file type. Use .csv or .xlsx")

    queue_items: list[QueueItem] = []
    skipped = 0
    for row in rows:
        mapped = _row_to_queue_item(row, niche=niche)
        if mapped is None:
            skipped += 1
            continue
        queue_items.append(mapped)

    generated, uploaded, rejected = _collect_status_stats(queue_items)
    result = QueueSyncResult(
        parsed_items=len(rows),
        upserted_items=len(queue_items),
        skipped_items=skipped,
        generated_items=generated,
        uploaded_items=uploaded,
        rejected_items=rejected,
    )
    with connect_db(db_path) as conn:
        upsert_queue_items(conn, queue_items)
        rebuild_publish_items(conn, niche)
        insert_sync_run(conn, str(path), niche, result)
    return result


def get_queue_summary(db_path: str | Path, niche: str) -> dict:
    with connect_db(db_path) as conn:
        return load_queue_summary(conn, niche)


def get_db_health(db_path: str | Path) -> dict:
    with connect_db(db_path) as conn:
        return check_db_health(conn)


def get_queue_stats(db_path: str | Path, niche: str | None, runs_limit: int = 10) -> dict:
    with connect_db(db_path) as conn:
        return load_queue_stats(conn, niche=niche, runs_limit=runs_limit)


def rebuild_publish_queue(db_path: str | Path, niche: str | None = None) -> dict:
    with connect_db(db_path) as conn:
        if niche:
            rebuilt = rebuild_publish_items(conn, niche=niche)
            return {"mode": "single", "niche": niche, "rebuilt_rows": rebuilt}
        rebuilt_by_niche = rebuild_publish_items_all(conn)
        return {
            "mode": "all",
            "niches": rebuilt_by_niche,
            "rebuilt_rows": sum(rebuilt_by_niche.values()),
        }


def export_n8n_ready_csv(
    db_path: str | Path,
    niche: str,
    output_path: str | Path,
    limit: int = 500,
    profile: str = "n8n_default",
    statuses: tuple[str, ...] = ("generated", "uploaded"),
) -> int:
    export_profiles = _load_export_profiles()
    if profile not in export_profiles:
        supported = ", ".join(sorted(export_profiles.keys()))
        raise ValueError(f"Unknown export profile '{profile}'. Supported: {supported}")
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with connect_db(db_path) as conn:
        rows = list_publish_items_for_n8n(conn, niche=niche, limit=limit, statuses=statuses)

    fieldnames = export_profiles[profile]
    with out.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})
    return len(rows)


def export_n8n_ready_json(
    db_path: str | Path,
    niche: str,
    output_path: str | Path,
    limit: int = 500,
    profile: str = "n8n_default",
    statuses: tuple[str, ...] = ("generated", "uploaded"),
) -> int:
    export_profiles = _load_export_profiles()
    if profile not in export_profiles:
        supported = ", ".join(sorted(export_profiles.keys()))
        raise ValueError(f"Unknown export profile '{profile}'. Supported: {supported}")
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with connect_db(db_path) as conn:
        rows = list_publish_items_for_n8n(conn, niche=niche, limit=limit, statuses=statuses)

    fieldnames = export_profiles[profile]
    payload = [{key: row.get(key, "") for key in fieldnames} for row in rows]
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(payload)
